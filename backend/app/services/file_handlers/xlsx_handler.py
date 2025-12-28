"""
XLSX handler using openpyxl.
"""
import logging
from typing import Dict, Any, List
from pathlib import Path

try:
    import openpyxl
    from openpyxl import load_workbook
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

from .base_handler import BaseFileHandler

logger = logging.getLogger(__name__)


class XLSXHandler(BaseFileHandler):
    """
    XLSX handler using openpyxl.
    """
    
    SUPPORTED_TYPES = ['xlsx', 'xls']
    
    def __init__(self):
        """Initialize the XLSX handler."""
        if not XLSX_AVAILABLE:
            raise ImportError(
                "openpyxl library not available. "
                "Install with: pip install openpyxl"
            )
    
    def can_handle(self, file_path: str, file_type: str) -> bool:
        """Check if file type is supported."""
        return file_type.lower() in ['xlsx', 'xls'] and XLSX_AVAILABLE
    
    def get_supported_types(self) -> List[str]:
        """Get list of supported file types."""
        return self.SUPPORTED_TYPES if XLSX_AVAILABLE else []
    
    def process(self, file_path: str) -> Dict[str, Any]:
        """
        Process XLSX file using openpyxl.
        
        Args:
            file_path: Path to file
            
        Returns:
            Dictionary with extracted text and metadata
        """
        if not self.validate_file(file_path):
            raise ValueError(f"Invalid file: {file_path}")
        
        try:
            logger.info(f"Processing XLSX with openpyxl: {file_path}")
            
            # Load workbook
            workbook = load_workbook(file_path, data_only=True)
            
            text_content = []
            metadata = {
                "sheet_count": len(workbook.sheetnames),
                "sheet_names": workbook.sheetnames,
                "has_data": False
            }
            
            # Process each sheet - CRITICAL: Process ALL sheets, not just the first one
            logger.info(f"Processing {len(workbook.sheetnames)} sheets: {workbook.sheetnames}")
            for sheet_idx, sheet_name in enumerate(workbook.sheetnames, 1):
                sheet = workbook[sheet_name]
                logger.info(f"Processing sheet {sheet_idx}/{len(workbook.sheetnames)}: '{sheet_name}'")
                text_content.append(f"=== Sheet {sheet_idx}: {sheet_name} ===")
                
                # Extract data from cells
                sheet_data = []
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                logger.info(f"  Sheet '{sheet_name}' has {max_row} rows x {max_col} columns")
                
                if max_row > 0 and max_col > 0:
                    metadata["has_data"] = True
                    
                    # Read all cells - process entire sheet
                    rows_processed = 0
                    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=False):
                        row_values = []
                        for cell in row:
                            if cell.value is not None:
                                # Convert cell value to string
                                cell_value = str(cell.value).strip()
                                if cell_value:
                                    row_values.append(cell_value)
                        
                        if row_values:
                            # Join row values with separator
                            sheet_data.append(" | ".join(row_values))
                            rows_processed += 1
                    
                    logger.info(f"  Extracted {rows_processed} rows from sheet '{sheet_name}'")
                    
                    if sheet_data:
                        text_content.extend(sheet_data)
                        text_content.append("")  # Empty line between sheets
                    else:
                        logger.warning(f"  No data extracted from sheet '{sheet_name}'")
                else:
                    logger.warning(f"  Sheet '{sheet_name}' appears empty (max_row={max_row}, max_col={max_col})")
            
            logger.info(f"✅ Processed all {len(workbook.sheetnames)} sheets from XLSX file")
            
            full_text = "\n".join(text_content)
            
            if not full_text.strip():
                logger.warning(f"No text extracted from XLSX: {file_path}")
                full_text = "[XLSX file - text extraction may be limited]"
            
            # Create simple chunks
            chunks = self._create_simple_chunks(full_text)
            
            return {
                "text": full_text,
                "metadata": metadata,
                "chunks": chunks
            }
            
        except Exception as e:
            logger.error(f"Error processing XLSX {file_path}: {str(e)}")
            raise RuntimeError(f"Failed to process XLSX: {str(e)}")
    
    def _create_simple_chunks(self, text: str, chunk_size: int = 1000, overlap: int = 200) -> List[Dict[str, Any]]:
        """
        Create simple text chunks.
        
        Args:
            text: Text to chunk
            chunk_size: Size of each chunk
            overlap: Overlap between chunks
            
        Returns:
            List of chunks with text and metadata
        """
        chunks = []
        start = 0
        text_length = len(text)
        chunk_index = 0
        
        while start < text_length:
            end = start + chunk_size
            chunk_text = text[start:end]
            
            # Try to break at line boundary (for spreadsheet data)
            if end < text_length:
                last_newline = chunk_text.rfind('\n')
                if last_newline > chunk_size * 0.5:
                    chunk_text = chunk_text[:last_newline + 1]
                    end = start + last_newline + 1
            
            chunks.append({
                "text": chunk_text.strip(),
                "metadata": {
                    "chunk_index": chunk_index,
                    "start_char": start,
                    "end_char": end,
                }
            })
            
            chunk_index += 1
            start = end - overlap
        
        return chunks

